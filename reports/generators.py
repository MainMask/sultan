from io import BytesIO
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.db.models import Sum, Count, Avg


HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
SUBHEADER_FILL = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
ALT_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)


def _header_row(ws, row, values):
    for i, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=i, value=val)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER


def _data_row(ws, row, values, alt=False):
    for i, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=i, value=val)
        cell.border = THIN_BORDER
        if alt:
            cell.fill = ALT_FILL


def generate_group_report(group_id, period_start, period_end, group_by='athlete') -> BytesIO:
    from core.models import Training, Group, Athlete

    group = Group.objects.select_related('trainer').get(pk=group_id)
    athletes = Athlete.objects.filter(group=group).order_by('full_name')

    trainings = Training.objects.filter(
        athlete__in=athletes,
        date__gte=period_start,
        date__lte=period_end,
    ).select_related('athlete', 'exercise').order_by('date', 'athlete__full_name')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Групповой отчёт'

    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f'Групповой отчёт: {group.name} | {period_start.strftime("%d.%m.%Y")} – {period_end.strftime("%d.%m.%Y")}'
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal='center')

    ws.append([])

    if group_by == 'date':
        _header_row(ws, 3, ['Дата', 'Атлет', 'Упражнение', 'Тип нагрузки', 'Подходы', 'Повторения', 'Вес (кг)', 'Длит. (мин)'])
        for i, t in enumerate(trainings, 4):
            _data_row(ws, i, [
                t.date.strftime('%d.%m.%Y'), t.athlete.full_name,
                t.exercise.name, t.exercise.get_load_type_display(),
                t.sets, t.reps, float(t.weight) if t.weight else None, t.duration,
            ], alt=i % 2 == 0)
    else:
        _header_row(ws, 3, ['Атлет', 'Дата', 'Упражнение', 'Тип нагрузки', 'Подходы', 'Повторения', 'Вес (кг)', 'Длит. (мин)'])
        for i, t in enumerate(trainings.order_by('athlete__full_name', 'date'), 4):
            _data_row(ws, i, [
                t.athlete.full_name, t.date.strftime('%d.%m.%Y'),
                t.exercise.name, t.exercise.get_load_type_display(),
                t.sets, t.reps, float(t.weight) if t.weight else None, t.duration,
            ], alt=i % 2 == 0)

    _auto_width(ws)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_athlete_report(athlete_id, period_start, period_end) -> BytesIO:
    from core.models import Training, Athlete

    athlete = Athlete.objects.select_related('group', 'trainer').get(pk=athlete_id)
    trainings = Training.objects.filter(
        athlete=athlete, date__gte=period_start, date__lte=period_end,
    ).select_related('exercise').order_by('date')

    wb = Workbook()

    # Лист 1 — все тренировки
    ws1 = wb.active
    ws1.title = 'Тренировки'
    ws1.merge_cells('A1:G1')
    ws1['A1'].value = f'{athlete.full_name} | {period_start.strftime("%d.%m.%Y")} – {period_end.strftime("%d.%m.%Y")}'
    ws1['A1'].font = Font(bold=True, size=13)
    ws1['A1'].alignment = Alignment(horizontal='center')
    ws1.append([])

    _header_row(ws1, 3, ['Дата', 'Упражнение', 'Тип нагрузки', 'Подходы', 'Повторения', 'Вес (кг)', 'Длит. (мин)'])
    for i, t in enumerate(trainings, 4):
        _data_row(ws1, i, [
            t.date.strftime('%d.%m.%Y'), t.exercise.name,
            t.exercise.get_load_type_display(), t.sets, t.reps,
            float(t.weight) if t.weight else None, t.duration,
        ], alt=i % 2 == 0)
    _auto_width(ws1)

    # Лист 2 — статистика по типу нагрузки
    ws2 = wb.create_sheet('Статистика по типам')
    _header_row(ws2, 1, ['Тип нагрузки', 'Кол-во тренировок', 'Сумм. подходы', 'Сумм. повторения', 'Макс. вес (кг)', 'Сумм. время (мин)'])

    from django.db.models import Max
    stats = trainings.values('exercise__load_type', 'exercise__load_type').annotate(
        cnt=Count('id'),
        total_sets=Sum('sets'),
        total_reps=Sum('reps'),
        max_weight=Max('weight'),
        total_duration=Sum('duration'),
    )
    load_type_map = dict(Training._meta.get_field('exercise').remote_field.model.load_type.field.choices) if False else {
        'strength': 'Силовая', 'cardio': 'Кардио', 'flexibility': 'Гибкость',
        'speed': 'Скоростная', 'endurance': 'Выносливость', 'technical': 'Техническая',
    }
    from core.models import LOAD_TYPE_CHOICES
    load_map = dict(LOAD_TYPE_CHOICES)

    for i, s in enumerate(stats, 2):
        _data_row(ws2, i, [
            load_map.get(s['exercise__load_type'], s['exercise__load_type']),
            s['cnt'], s['total_sets'], s['total_reps'],
            float(s['max_weight']) if s['max_weight'] else None,
            s['total_duration'],
        ], alt=i % 2 == 0)
    _auto_width(ws2)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_attendance_report(group_id, period_start, period_end) -> BytesIO:
    from core.models import Training, Group, Athlete

    group = Group.objects.get(pk=group_id)
    athletes = Athlete.objects.filter(group=group).order_by('full_name')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Посещаемость'

    ws.merge_cells('A1:E1')
    ws['A1'].value = f'Посещаемость группы: {group.name} | {period_start.strftime("%d.%m.%Y")} – {period_end.strftime("%d.%m.%Y")}'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.append([])

    _header_row(ws, 3, ['Атлет', 'Кол-во тренировок', 'Сумм. время (мин)', 'Человеко-часы', 'Общий объём (подх.×повт.)'])

    for i, athlete in enumerate(athletes, 4):
        stats = Training.objects.filter(
            athlete=athlete, date__gte=period_start, date__lte=period_end,
        ).aggregate(
            cnt=Count('id'),
            total_duration=Sum('duration'),
            total_volume=Sum('sets') if False else Count('id'),
        )
        raw = Training.objects.filter(
            athlete=athlete, date__gte=period_start, date__lte=period_end,
        )
        cnt = raw.count()
        total_dur = raw.aggregate(s=Sum('duration'))['s'] or 0
        person_hours = round(total_dur / 60, 2)
        volume = sum((t.sets or 0) * (t.reps or 1) for t in raw)

        _data_row(ws, i, [athlete.full_name, cnt, total_dur, person_hours, volume], alt=i % 2 == 0)

    _auto_width(ws)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_personal_report(athlete_id, period_start, period_end, group_by='week', exercise_type='') -> BytesIO:
    from core.models import Training, Athlete, LOAD_TYPE_CHOICES

    athlete = Athlete.objects.get(pk=athlete_id)
    qs = Training.objects.filter(
        athlete=athlete, date__gte=period_start, date__lte=period_end,
    ).select_related('exercise').order_by('date')

    if exercise_type:
        qs = qs.filter(exercise__load_type=exercise_type)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Личный отчёт'

    ws.merge_cells('A1:G1')
    ws['A1'].value = f'Личный отчёт: {athlete.full_name} | {period_start.strftime("%d.%m.%Y")} – {period_end.strftime("%d.%m.%Y")}'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.append([])

    if group_by == 'week':
        _header_row(ws, 3, ['Неделя', 'Упражнение', 'Тип', 'Подходы', 'Повторения', 'Вес (кг)', 'Длит. (мин)'])
        row = 4
        for t in qs:
            week_start = t.date - timedelta(days=t.date.weekday())
            week_label = f'Нед. {week_start.strftime("%d.%m.%Y")}'
            _data_row(ws, row, [
                week_label, t.exercise.name, t.exercise.get_load_type_display(),
                t.sets, t.reps, float(t.weight) if t.weight else None, t.duration,
            ], alt=row % 2 == 0)
            row += 1
    else:
        _header_row(ws, 3, ['Месяц', 'Упражнение', 'Тип', 'Подходы', 'Повторения', 'Вес (кг)', 'Длит. (мин)'])
        row = 4
        for t in qs:
            month_label = t.date.strftime('%B %Y')
            _data_row(ws, row, [
                month_label, t.exercise.name, t.exercise.get_load_type_display(),
                t.sets, t.reps, float(t.weight) if t.weight else None, t.duration,
            ], alt=row % 2 == 0)
            row += 1

    _auto_width(ws)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
